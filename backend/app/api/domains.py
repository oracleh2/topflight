import os
import requests
import tempfile
from operator import and_
from typing import List
from docx import Document
import openpyxl
from io import StringIO, BytesIO

from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app import logger, UserDomain, UserKeyword
from app.core.user_service import UserService
from app.database import get_session
from app.dependencies import get_current_user, require_api_key
from app.models import User, Region, YandexRegion, DeviceType
from app.schemas.domain import (
    DomainAdd,
    DomainResponse,
    KeywordAdd,
    KeywordResponse,
    RegionResponse,
    BulkKeywordsAdd,
    WordFileLoad,
    ExcelFileLoad,
    TextFileLoad,
    KeywordUpdate,
    BulkDeleteKeywords,
    BulkEditKeywords,
)

router = APIRouter(prefix="/domains", tags=["Domains"])


@router.get("/regions", response_model=List[RegionResponse])
async def get_regions(
    limit: int = Query(
        100, ge=1, le=1000, description="Максимальное количество результатов"
    ),
    session: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Получение списка доступных регионов"""
    result = await session.execute(
        select(YandexRegion)
        .where(YandexRegion.is_active == True)
        .order_by(YandexRegion.region_name)
        .limit(limit)
    )
    regions = result.scalars().all()

    return [
        RegionResponse(
            id=str(region.id),
            code=region.region_code,
            name=region.display_name or region.region_name,
            country_code=region.country_code,
            region_type=region.region_type,
        )
        for region in regions
    ]


@router.get("/regions/search", response_model=List[RegionResponse])
async def search_regions(
    q: str = Query(..., min_length=2, max_length=100, description="Поисковый запрос"),
    limit: int = Query(
        20, ge=1, le=50, description="Максимальное количество результатов"
    ),
    session: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Поиск регионов по названию или коду"""
    try:
        # Используем улучшенный метод поиска с приоритизацией
        regions = await YandexRegion.search_with_region_type_priority(session, q, limit)

        return [
            RegionResponse(
                id=str(region.id),
                code=region.region_code,
                name=region.display_name or region.region_name,
                country_code=region.country_code,
                region_type=region.region_type,  # Добавляем тип региона
            )
            for region in regions
        ]
    except Exception as e:
        logger.error(f"Error searching regions: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Ошибка поиска регионов",
        )


@router.post("/", response_model=dict)
async def add_domain(
    domain_data: DomainAdd,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Добавление домена"""
    user_service = UserService(session)
    result = await user_service.add_domain(
        user_id=str(current_user.id), domain=domain_data.domain
    )

    if not result["success"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=result["errors"]
        )

    return {
        "success": True,
        "message": "Домен успешно добавлен",
        "domain_id": result["domain_id"],
        "domain": result["domain"],
    }


@router.get("/", response_model=List[DomainResponse])
async def get_user_domains(
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Получение доменов пользователя"""
    user_service = UserService(session)
    domains = await user_service.get_user_domains(str(current_user.id))

    return [DomainResponse(**domain) for domain in domains]


@router.post("/{domain_id}/keywords", response_model=dict)
async def add_keyword(
    domain_id: str,
    keyword_data: KeywordAdd,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Добавление ключевого слова"""
    user_service = UserService(session)
    result = await user_service.add_keyword(
        user_id=str(current_user.id),
        domain_id=domain_id,
        keyword=keyword_data.keyword,
        region_id=keyword_data.region_id,
        device_type=keyword_data.device_type,
        check_frequency=keyword_data.check_frequency,
        is_active=keyword_data.is_active,
    )

    if not result["success"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=result["errors"]
        )

    return {
        "success": True,
        "message": "Ключевое слово успешно добавлено",
        "keyword_id": result["keyword_id"],
    }


@router.get("/{domain_id}/keywords", response_model=List[KeywordResponse])
async def get_domain_keywords(
    domain_id: str,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Получение ключевых слов домена"""
    user_service = UserService(session)
    keywords = await user_service.get_domain_keywords(
        user_id=str(current_user.id), domain_id=domain_id
    )

    return [KeywordResponse(**keyword) for keyword in keywords]


@router.get("/{domain_id}/proxy-settings")
async def get_domain_proxy_settings(
    domain_id: str,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    # Загрузка настроек прокси для домена
    pass


@router.put("/{domain_id}/proxy-settings")
async def update_domain_proxy_settings(
    domain_id: str,
    settings: dict,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    # Сохранение настроек прокси для домена
    pass


@router.get("/proxies/domain/{domain_id}/stats")
async def get_domain_proxy_stats(
    domain_id: str,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    # Статистика прокси домена
    pass


@router.delete("/keywords/bulk", response_model=dict)
async def bulk_delete_keywords(
    delete_data: BulkDeleteKeywords,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Массовое удаление ключевых слов"""
    try:
        # Получаем ключевые слова для удаления
        keywords = await session.execute(
            select(UserKeyword).where(
                UserKeyword.id.in_(delete_data.keyword_ids),
                UserKeyword.user_id == current_user.id,
            )
        )
        keywords_to_delete = keywords.scalars().all()

        if not keywords_to_delete:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Ключевые слова не найдены",
            )

        deleted_count = len(keywords_to_delete)

        # Удаляем все ключевые слова
        for keyword in keywords_to_delete:
            await session.delete(keyword)

        await session.commit()

        logger.info(f"Bulk deleted {deleted_count} keywords by user {current_user.id}")

        return {
            "success": True,
            "message": f"Удалено {deleted_count} ключевых слов",
            "deleted_count": deleted_count,
        }

    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        logger.error(f"Failed to bulk delete keywords: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Ошибка при массовом удалении ключевых слов",
        )


@router.put("/keywords/{keyword_id}", response_model=dict)
async def update_keyword(
    keyword_id: str,
    keyword_data: KeywordUpdate,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Обновление ключевого слова"""
    try:
        # Проверяем существование ключевого слова и принадлежность пользователю
        keyword = await session.execute(
            select(UserKeyword).where(
                UserKeyword.id == keyword_id, UserKeyword.user_id == current_user.id
            )
        )
        keyword_obj = keyword.scalar_one_or_none()

        if not keyword_obj:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Ключевое слово не найдено",
            )

        # Проверяем существование региона
        region = await session.get(YandexRegion, keyword_data.region_id)
        if not region:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Регион не найден"
            )

        # Проверяем на дубликаты (исключая текущее ключевое слово)
        existing_keyword = await session.execute(
            select(UserKeyword).where(
                UserKeyword.user_id == current_user.id,
                UserKeyword.domain_id == keyword_obj.domain_id,
                UserKeyword.keyword == keyword_data.keyword,
                UserKeyword.region_id == keyword_data.region_id,
                UserKeyword.device_type == DeviceType(keyword_data.device_type),
                UserKeyword.id != keyword_id,
            )
        )

        if existing_keyword.scalar_one_or_none():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ключевое слово уже существует для этого домена и региона",
            )

        # Обновляем ключевое слово
        keyword_obj.keyword = keyword_data.keyword
        keyword_obj.region_id = keyword_data.region_id
        keyword_obj.device_type = DeviceType(keyword_data.device_type)
        keyword_obj.check_frequency = keyword_data.check_frequency
        keyword_obj.is_active = keyword_data.is_active

        await session.commit()

        logger.info(f"Keyword updated: {keyword_id} by user {current_user.id}")

        return {
            "success": True,
            "message": "Ключевое слово успешно обновлено",
            "keyword_id": keyword_id,
        }

    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        logger.error(f"Failed to update keyword {keyword_id}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Ошибка при обновлении ключевого слова",
        )


@router.delete("/keywords/{keyword_id}", response_model=dict)
async def delete_keyword(
    keyword_id: str,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Удаление ключевого слова"""
    try:
        # Проверяем существование ключевого слова и принадлежность пользователю
        keyword = await session.execute(
            select(UserKeyword).where(
                UserKeyword.id == keyword_id, UserKeyword.user_id == current_user.id
            )
        )
        keyword_obj = keyword.scalar_one_or_none()

        if not keyword_obj:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Ключевое слово не найдено",
            )

        # Удаляем ключевое слово
        await session.delete(keyword_obj)
        await session.commit()

        logger.info(f"Keyword deleted: {keyword_id} by user {current_user.id}")

        return {"success": True, "message": "Ключевое слово успешно удалено"}

    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        logger.error(f"Failed to delete keyword {keyword_id}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Ошибка при удалении ключевого слова",
        )


# API endpoints для работы через API ключ
@router.post("/api/add", response_model=dict)
async def api_add_domain(
    domain_data: DomainAdd,
    api_key: str = Query(..., description="API ключ"),
    session: AsyncSession = Depends(get_session),
):
    """Добавление домена через API"""
    current_user = await require_api_key(api_key=api_key, session=session)

    user_service = UserService(session)
    result = await user_service.add_domain(
        user_id=str(current_user.id), domain=domain_data.domain
    )

    if not result["success"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=result["errors"]
        )

    return {
        "success": True,
        "domain_id": result["domain_id"],
        "domain": result["domain"],
    }


@router.get("/api/list", response_model=List[DomainResponse])
async def api_get_domains(
    api_key: str = Query(..., description="API ключ"),
    session: AsyncSession = Depends(get_session),
):
    """Получение доменов через API"""
    current_user = await require_api_key(api_key=api_key, session=session)

    user_service = UserService(session)
    domains = await user_service.get_user_domains(str(current_user.id))

    return [DomainResponse(**domain) for domain in domains]


@router.post("/load-keywords/text-file", response_model=dict)
async def load_keywords_from_text_file(
    file_data: TextFileLoad,
    current_user: User = Depends(get_current_user),
):
    """Загрузка ключевых слов из текстового файла"""
    try:
        # Загружаем файл
        content = download_file(file_data.url)

        # Определяем кодировку и декодируем
        text_content = content.decode("utf-8")

        # Извлекаем ключевые слова
        keywords = extract_keywords_from_text(text_content)

        if not keywords:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="В файле не найдено ключевых слов",
            )

        return {"success": True, "keywords": keywords, "count": len(keywords)}

    except UnicodeDecodeError:
        # Пробуем другие кодировки
        try:
            text_content = content.decode("cp1251")
            keywords = extract_keywords_from_text(text_content)

            return {"success": True, "keywords": keywords, "count": len(keywords)}
        except:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Не удалось определить кодировку файла",
            )
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/load-keywords/excel", response_model=dict)
async def load_keywords_from_excel(
    file_data: ExcelFileLoad,
    current_user: User = Depends(get_current_user),
):
    """Загрузка ключевых слов из Excel файла"""
    try:
        # Загружаем файл
        content = download_file(file_data.url)

        # Извлекаем ключевые слова
        keywords = extract_keywords_from_excel(
            content, file_data.sheet, file_data.start_row
        )

        if not keywords:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="В файле не найдено ключевых слов",
            )

        return {"success": True, "keywords": keywords, "count": len(keywords)}

    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/load-keywords/word", response_model=dict)
async def load_keywords_from_word(
    file_data: WordFileLoad,
    current_user: User = Depends(get_current_user),
):
    """Загрузка ключевых слов из Word документа"""
    try:
        # Загружаем файл
        content = download_file(file_data.url)

        # Извлекаем ключевые слова
        keywords = extract_keywords_from_word(content)

        if not keywords:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="В документе не найдено ключевых слов",
            )

        return {"success": True, "keywords": keywords, "count": len(keywords)}

    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))


@router.post("/{domain_id}/keywords/bulk", response_model=dict)
async def add_bulk_keywords(
    domain_id: str,
    keywords_data: BulkKeywordsAdd,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Массовое добавление ключевых слов"""
    user_service = UserService(session)

    # Проверяем принадлежность домена пользователю
    domain = await session.execute(
        select(UserDomain).where(
            and_(UserDomain.id == domain_id, UserDomain.user_id == current_user.id)
        )
    )
    domain_obj = domain.scalar_one_or_none()
    if not domain_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Домен не найден"
        )

    # Проверяем существование региона
    region = await session.get(YandexRegion, keywords_data.region_id)
    if not region:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Регион не найден"
        )

    # Статистика добавления
    stats = {
        "total_keywords": len(keywords_data.keywords),
        "added_keywords": 0,
        "skipped_keywords": 0,
        "errors": [],
    }

    added_keyword_ids = []
    skipped_keywords = []

    try:
        for keyword in keywords_data.keywords:
            try:
                # Проверяем, существует ли уже такое ключевое слово
                existing_keyword = await session.execute(
                    select(UserKeyword).where(
                        UserKeyword.user_id == current_user.id,
                        UserKeyword.domain_id == domain_id,
                        UserKeyword.keyword == keyword,
                        UserKeyword.region_id == keywords_data.region_id,
                        UserKeyword.device_type
                        == DeviceType(keywords_data.device_type),
                    )
                )

                if existing_keyword.scalar_one_or_none():
                    skipped_keywords.append(keyword)
                    stats["skipped_keywords"] += 1
                    continue

                # Добавляем новое ключевое слово
                user_keyword = UserKeyword(
                    user_id=current_user.id,
                    domain_id=domain_id,
                    keyword=keyword,
                    region_id=keywords_data.region_id,
                    device_type=DeviceType(keywords_data.device_type),
                    check_frequency=keywords_data.check_frequency,
                    is_active=True,
                )

                session.add(user_keyword)
                await session.flush()  # Получаем ID без коммита

                added_keyword_ids.append(str(user_keyword.id))
                stats["added_keywords"] += 1

            except Exception as e:
                stats["errors"].append(f"Ошибка при добавлении '{keyword}': {str(e)}")
                continue

        # Коммитим все изменения
        await session.commit()

        logger.info(
            f"Bulk keywords added for user {current_user.id}, domain {domain_id}, stats: {stats}"
        )

        return {
            "success": True,
            "message": f"Добавлено {stats['added_keywords']} ключевых слов",
            "stats": stats,
            "added_keyword_ids": added_keyword_ids,
            "skipped_keywords": skipped_keywords,
        }

    except Exception as e:

        await session.rollback()

        logger.error(
            f"Failed to add bulk keywords for user {current_user.id}, domain {domain_id}, error: {str(e)}"
        )

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Ошибка при добавлении ключевых слов",
        )


@router.post("/{domain_id}/keywords/bulk-edit", response_model=dict)
async def bulk_edit_keywords(
    domain_id: str,
    edit_data: BulkEditKeywords,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Массовое редактирование ключевых слов"""
    try:
        # Проверяем принадлежность домена пользователю
        domain = await session.execute(
            select(UserDomain).where(
                UserDomain.id == domain_id, UserDomain.user_id == current_user.id
            )
        )
        domain_obj = domain.scalar_one_or_none()
        if not domain_obj:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Домен не найден"
            )

        stats = {"added_keywords": 0, "removed_keywords": 0, "errors": []}

        # Удаляем ключевые слова
        if edit_data.removed_keywords:
            keywords_to_remove = await session.execute(
                select(UserKeyword).where(
                    UserKeyword.id.in_(edit_data.removed_keywords),
                    UserKeyword.user_id == current_user.id,
                    UserKeyword.domain_id == domain_id,
                )
            )
            keywords_to_remove_list = keywords_to_remove.scalars().all()

            for keyword in keywords_to_remove_list:
                await session.delete(keyword)
                stats["removed_keywords"] += 1

        # Добавляем новые ключевые слова
        if edit_data.added_keywords and edit_data.new_keywords_settings:
            settings = edit_data.new_keywords_settings

            # Проверяем регион
            region = await session.get(YandexRegion, settings.get("region_id"))
            if not region:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND, detail="Регион не найден"
                )

            for keyword_text in edit_data.added_keywords:
                try:
                    # Проверяем на дубликаты
                    existing_keyword = await session.execute(
                        select(UserKeyword).where(
                            UserKeyword.user_id == current_user.id,
                            UserKeyword.domain_id == domain_id,
                            UserKeyword.keyword == keyword_text,
                            UserKeyword.region_id == settings.get("region_id"),
                            UserKeyword.device_type
                            == DeviceType(settings.get("device_type", "desktop")),
                        )
                    )

                    if existing_keyword.scalar_one_or_none():
                        stats["errors"].append(
                            f"Ключевое слово '{keyword_text}' уже существует"
                        )
                        continue

                    # Добавляем ключевое слово
                    new_keyword = UserKeyword(
                        user_id=current_user.id,
                        domain_id=domain_id,
                        keyword=keyword_text,
                        region_id=settings.get("region_id"),
                        device_type=DeviceType(settings.get("device_type", "desktop")),
                        check_frequency=settings.get("check_frequency", "daily"),
                        is_active=settings.get("is_active", True),
                    )

                    session.add(new_keyword)
                    stats["added_keywords"] += 1

                except Exception as e:
                    stats["errors"].append(
                        f"Ошибка при добавлении '{keyword_text}': {str(e)}"
                    )

        await session.commit()

        logger.info(
            f"Bulk edit completed for domain {domain_id} by user {current_user.id}: {stats}"
        )

        return {
            "success": True,
            "message": f"Добавлено {stats['added_keywords']}, удалено {stats['removed_keywords']} ключевых слов",
            "stats": stats,
        }

    except HTTPException:
        raise
    except Exception as e:
        await session.rollback()
        logger.error(f"Failed to bulk edit keywords for domain {domain_id}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Ошибка при массовом редактировании ключевых слов",
        )


# Вспомогательные функции
def download_file(url: str, max_size: int = 10 * 1024 * 1024) -> bytes:
    """Загружает файл по URL с ограничением размера"""
    try:
        response = requests.get(url, stream=True, timeout=30)
        response.raise_for_status()

        # Проверяем размер файла
        content_length = response.headers.get("content-length")
        if content_length and int(content_length) > max_size:
            raise ValueError(f"Файл слишком большой (максимум {max_size} байт)")

        # Загружаем файл по частям
        content = b""
        for chunk in response.iter_content(chunk_size=8192):
            content += chunk
            if len(content) > max_size:
                raise ValueError(f"Файл слишком большой (максимум {max_size} байт)")

        return content
    except requests.RequestException as e:
        raise ValueError(f"Ошибка загрузки файла: {str(e)}")


def extract_keywords_from_text(content: str) -> List[str]:
    """Извлекает ключевые слова из текста"""
    lines = content.split("\n")
    keywords = []

    for line in lines:
        line = line.strip()
        if line and len(line) <= 500:  # Ограничение длины ключевого слова
            keywords.append(line)

    return keywords


def extract_keywords_from_excel(
    content: bytes, sheet: int = 1, start_row: int = 1
) -> List[str]:
    """Извлекает ключевые слова из Excel файла"""
    try:
        # Создаем временный файл
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            tmp_file.write(content)
            tmp_file_path = tmp_file.name

        try:
            # Читаем Excel файл
            workbook = openpyxl.load_workbook(tmp_file_path)

            # Получаем лист
            if sheet > len(workbook.sheetnames):
                raise ValueError(
                    f"Лист {sheet} не найден. Доступно листов: {len(workbook.sheetnames)}"
                )

            worksheet = workbook.worksheets[sheet - 1]

            keywords = []
            for row in worksheet.iter_rows(
                min_row=start_row, max_col=1, values_only=True
            ):
                if row[0] and isinstance(row[0], str):
                    keyword = str(row[0]).strip()
                    if keyword and len(keyword) <= 500:
                        keywords.append(keyword)

            return keywords

        finally:
            # Удаляем временный файл
            os.unlink(tmp_file_path)

    except Exception as e:
        raise ValueError(f"Ошибка обработки Excel файла: {str(e)}")


def extract_keywords_from_word(content: bytes) -> List[str]:
    """Извлекает ключевые слова из Word документа"""
    try:
        # Создаем временный файл
        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp_file:
            tmp_file.write(content)
            tmp_file_path = tmp_file.name

        try:
            # Читаем Word документ
            doc = Document(tmp_file_path)

            keywords = []
            for paragraph in doc.paragraphs:
                text = paragraph.text.strip()
                if text and len(text) <= 500:
                    keywords.append(text)

            return keywords

        finally:
            # Удаляем временный файл
            os.unlink(tmp_file_path)

    except Exception as e:
        raise ValueError(f"Ошибка обработки Word документа: {str(e)}")


@router.delete("/{domain_id}", response_model=dict)
async def delete_domain(
    domain_id: str,
    current_user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Удаление домена"""
    user_service = UserService(session)
    result = await user_service.delete_domain(
        user_id=str(current_user.id), domain_id=domain_id
    )

    if not result["success"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=result["errors"]
        )

    return {"success": True, "message": "Домен успешно удален"}
